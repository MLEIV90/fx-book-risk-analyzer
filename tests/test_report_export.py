"""
Tests for the Excel report export.

Verifies the polish pass on report_export.py: proper number-format masks are
applied (not by rounding the underlying value -- the cell keeps full
precision), no truncating column widths, and the workbook opens cleanly.
"""
import io

import openpyxl

from report_export import build_excel_report


def _sample_report():
    summary = {
        "notional": 8129.272306733868,
        "book_value": 20026.42021618709,
        "confidence": 0.99,
        "var_parametric": 12345.6789,
        "var_historical": 23456.789,
        "var_montecarlo": 34567.891,
        "var_ewma": 45678.912,
        "var_student_t": 56789.123,
        "expected_shortfall": 67891.234,
        "limits": [{"label": "VaR", "status": "OK", "ok": True},
                   {"label": "Net exposure EUR", "status": "BREACH", "ok": False}],
    }
    book_rows = [
        {"pair": "EUR/USD", "side": "Provider buys EUR", "notional": 1_000_000.0,
         "rate": 1.150727345, "tenor": 90, "mtm": 8129.272306733868},
        {"pair": "EUR/GBP", "side": "Provider sells EUR", "notional": 500_000.0,
         "rate": 0.854021, "tenor": 120, "mtm": -2345.6789},
    ]
    risk_detail = {
        "dv01_by_ccy": {"EUR": -10.4567, "USD": 34.5321},
        "dv01_by_tenor": {"0-3m": 0.1274, "3-6m": 12.5},
        "liquidity": 20026.42021618709,
        "stress": {"Brexit referendum (Jun 2016)": -12345.6789},
    }
    xlsx_bytes = build_excel_report(book_rows=book_rows, summary=summary,
                                    risk_detail=risk_detail)
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes)), summary, book_rows, risk_detail


def test_workbook_opens_with_expected_sheets():
    """The workbook must open cleanly (not corrupt) with all three sheets."""
    wb, *_ = _sample_report()
    assert wb.sheetnames == ["Summary", "Positions", "Risk detail"]


def test_values_are_full_precision_not_rounded():
    """
    Formatting must be purely a DISPLAY change: the underlying cell value
    must still be the exact float passed in, not a rounded copy.
    """
    wb, summary, book_rows, risk_detail = _sample_report()
    s = wb["Summary"]
    found_notional = any(cell.value == summary["notional"]
                         for row in s.iter_rows() for cell in row)
    found_book_value = any(cell.value == summary["book_value"]
                           for row in s.iter_rows() for cell in row)
    assert found_notional
    assert found_book_value

    p = wb["Positions"]
    found_rate = any(cell.value == book_rows[0]["rate"]
                     for row in p.iter_rows() for cell in row)
    assert found_rate


def test_currency_cells_use_thousands_zero_decimal_format():
    """USD amounts must display with thousands separators and no decimals."""
    wb, summary, *_ = _sample_report()
    s = wb["Summary"]
    cell = next(c for row in s.iter_rows() for c in row
               if c.value == summary["notional"])
    assert cell.number_format == "#,##0;(#,##0)"


def test_dv01_cells_use_two_decimal_format():
    wb, *_ , risk_detail = _sample_report()
    d = wb["Risk detail"]
    cell = next(c for row in d.iter_rows() for c in row
               if c.value == risk_detail["dv01_by_ccy"]["EUR"])
    assert cell.number_format == "#,##0.00;(#,##0.00)"


def test_rate_cells_use_four_decimal_format():
    wb, _, book_rows, _ = _sample_report()
    p = wb["Positions"]
    cell = next(c for row in p.iter_rows() for c in row
               if c.value == book_rows[0]["rate"])
    assert cell.number_format == "0.0000"


def test_tenor_cells_use_integer_format():
    """Regression: tenor previously had no number format (plain General)."""
    wb, _, book_rows, _ = _sample_report()
    p = wb["Positions"]
    cell = next(c for row in p.iter_rows() for c in row
               if c.value == book_rows[0]["tenor"])
    assert cell.number_format == "0"


def test_confidence_cell_uses_percentage_format():
    wb, summary, *_ = _sample_report()
    s = wb["Summary"]
    cell = next(c for row in s.iter_rows() for c in row
               if c.value == summary["confidence"])
    assert cell.number_format == "0.0%"


def test_risk_detail_column_a_wide_enough_for_longest_header():
    """
    Regression: 'Key-rate DV01 by tenor (USD per 1bp)' (36 chars) previously
    didn't fit in a 28-wide column.
    """
    wb, *_ = _sample_report()
    d = wb["Risk detail"]
    width = d.column_dimensions["A"].width
    assert width >= 36


def test_positions_side_column_wide_enough_for_longest_value():
    """Regression: 'Provider sells EUR' (18 chars) previously didn't fit in
    a 16-wide column."""
    wb, *_ = _sample_report()
    p = wb["Positions"]
    width = p.column_dimensions["B"].width
    assert width >= 18


def test_summary_footer_states_governing_var():
    """The report must be self-explanatory about which VaR governs the
    limit check, without needing the app alongside it."""
    wb, *_ = _sample_report()
    s = wb["Summary"]
    text = "\n".join(str(c.value) for row in s.iter_rows() for c in row if c.value)
    assert "GOVERNING VaR" in text
    assert "max(historical, Student-t)" in text
